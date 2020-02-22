__author__ = 'Arshiya28'

from openpyxl import load_workbook
import os,sys

def createIncomeStatement(inputFile,inputDate):
    wbFile =  load_workbook(filename = inputFile, use_iterators = True)
    sheet_ranges = wbFile['Sheet1']
    date = sheet_ranges['J3'].value
    suffix = date.replace("-","")
    outputFile = open('C:\Users\Arshiya28\Desktop\MBA\Project\Software\Step-2\Output\IncomeStatement_'+inputDate,"wb")
    outputFile.write("================================================================================================\n")
    outputFile.write("                                    INCOME STATEMENT\n")
    outputFile.write("================================================================================================\n")
    outputFile.write("Company Name: "+sheet_ranges['D3'].value+"\n")
    outputFile.write("For the Year Ending "+date+"\n")
    outputFile.write("================================================================================================\n")
    outputFile.write("REVENUE\n")
    outputFile.write("========\n")
    ws = wbFile.get_sheet_by_name(name = 'Sheet1')
    x = 7
    y = ""
    flag = 0
    total_revenue = 0.00
    trailing_spaces = 0
    for row in ws.iter_rows('R7:R20'):
        for cell in row:
            if str(cell.value) != "None":
                if cell.value == "TOTAL_REVENUE":
                    outputFile.write("-------------------------------------------------------------------------------------------------\n")
                    outputFile.write("    Total Revenue                                                               $"+"%.2f" % total_revenue+"\n")
                    outputFile.write("-------------------------------------------------------------------------------------------------\n")
                    break
                else:
                    y = "T"+str(x)
                    line = "    " + str(cell.value)
                    if len(line) < 61:
                        trailing_spaces = 61-len(line)
                        line = line + " " * trailing_spaces
                    if str(ws.cell(y).value) != "None":
                        line = line + "$"+"%.2f" % float(ws.cell(y).value)
                    else:
                        line =line + "$0.00"
                    if cell.value == "Loss":
                        total_revenue =total_revenue - (ws.cell(y).value)
                    else:
                        total_revenue =total_revenue + (ws.cell(y).value)
                    outputFile.write(line+"\n")
        x = x + 1
    p = 7
    q = ""
    gross_income = total_revenue
    for row in ws.iter_rows('W7:W2000'):
        for cell in row:
            if str(cell.value) != "None":
                if str(cell.value) == "Cost_of_Goods_Sold":
                    flag = 1
                    q = "Y"+str(p)
                    if str(ws.cell(q).value) != "None":
                        gross_income = gross_income - (ws.cell(q).value)
                        outputFile.write("    Cost of Goods Sold                                                          $"+"%.2f" % ws.cell(q).value+"\n")
                    else:
                        outputFile.write("    Cost of Goods Sold                                                          $0.00\n")
                elif str(cell.value) == "TOTAL_EXPENSES":
                    break
        p = p + 1
    if flag != 1:
        outputFile.write("    Cost of Goods Sold                                                          $0.00\n")

    outputFile.write("-------------------------------------------------------------------------------------------------\n")
    outputFile.write("    Gross Income                                                                $"+ "%.2f" % gross_income +"\n")
    outputFile.write("-------------------------------------------------------------------------------------------------\n")
    total_expenses = 0.00
    i = 7
    j = ""
    trailing_spaces = 0
    interest_expense = 0.00
    income_tax_expense = 0.00
    interest = []
    outputFile.write("OPERATING EXPENSES\n")
    outputFile.write("===================\n")
    for row in ws.iter_rows('W7:W2000'):
        for cell in row:
            if str(cell.value) != "None":
                if str(cell.value) == "TOTAL_EXPENSES":
                    outputFile.write("-------------------------------------------------------------------------------------------------\n")
                    outputFile.write("    Total Expenses                                                              $"+"%.2f" % total_expenses+"\n")
                    outputFile.write("-------------------------------------------------------------------------------------------------\n")
                    break
                elif  "Interest_Expense" in str(cell.value):
                    j = "Y"+str(i)
                    if str(ws.cell(j).value) != "None":
                        interest_expense = interest_expense + ws.cell(j).value
                        interest.append(cell.value+"#"+str(ws.cell(j).value))
                    else:
                        interest_expense = 0.00
                        interest.append(cell.value+"#0.00")
                elif str(cell.value) == "Income_Tax_Expense":
                    j = "Y"+str(i)
                    if str(ws.cell(j).value) != "None":
                        income_tax_expense = ws.cell(j).value
                    else:
                        income_tax_expense = 0.00
                elif str(cell.value) == "Cost_of_Goods_Sold":
                    continue
                else:
                    j = "Y"+str(i)
                    line = "    " + str(cell.value)
                    if len(line) < 61:
                        trailing_spaces = 61-len(line)
                        line = line + " " * trailing_spaces
                    if str(ws.cell(j).value) != "None":
                        line = line + "$"+"%.2f" % ws.cell(j).value
                    else:
                        line =line + "$0.00"
                    total_expenses =total_expenses + ws.cell(j).value
                    outputFile.write(line+"\n")
        i = i + 1
    operating_profit = gross_income - total_expenses
    outputFile.write("    OPERATING PROFIT (Earning Before Income and Tax)                            $"+"%.2f" % operating_profit +"\n")
    outputFile.write("-------------------------------------------------------------------------------------------------\n")
    outputFile.write("NON-OPERATING EXPENSES\n")
    outputFile.write("======================\n")
    outputFile.write("    INTEREST EXPENSES\n")
    outputFile.write("    =================\n")

    for element in interest:
        line = ""
        line = "    "+element.split("#")[0]
        if len(line) < 61:
            trailing_spaces = 61 - len(line)
            line = line + " " * trailing_spaces
        outputFile.write(line+"$"+"%.2f" %float(element.split("#")[1])+"\n")
    netIncomeAfterInterest = operating_profit - interest_expense
    outputFile.write("-------------------------------------------------------------------------------------------------\n")
    outputFile.write("    TOTAL INTEREST EXPENSES                                                     $"+"%.2f" % interest_expense + "\n")
    outputFile.write("-------------------------------------------------------------------------------------------------\n")
    outputFile.write("    NET INCOME After Interest                                                   $"+"%.2f" % netIncomeAfterInterest + "\n")
    outputFile.write("-------------------------------------------------------------------------------------------------\n")
    outputFile.write("    INCOME TAX EXPENSE                                                          $"+"%.2f" % income_tax_expense + "\n")
    outputFile.write("-------------------------------------------------------------------------------------------------\n")
    netIncomeAfterTax = netIncomeAfterInterest - income_tax_expense
    outputFile.write("    NET INCOME After TAX                                                        $"+"%.2f" % netIncomeAfterTax + "\n")
    outputFile.write("=================================================================================================\n")
    if netIncomeAfterTax < 0:
        outputFile.write("    NET LOSS                                                                    $"+"%.2f" % netIncomeAfterTax + "\n")
    elif netIncomeAfterTax > 0:
        outputFile.write("    NET PROFIT                                                                  $"+"%.2f" % netIncomeAfterTax + "\n")
    elif netIncomeAfterTax == 0:
        outputFile.write("    BREAK EVEN POINT                                                            $0.00\n")
    outputFile.write("=================================================================================================\n")
    outputFile.close()
    return netIncomeAfterTax

def createCashFlowStatements(inFile1,infile2,netIncome,dividends,inputDate):

    addition = {}
    substraction = {}
    beginingCash = 0
    # First Balance Sheet
    wbFile1 =  load_workbook(filename = inFile1, use_iterators = True)
    sheet_ranges1 = wbFile1['Sheet1']
    #Second Balance Sheet
    wbFile2 =  load_workbook(filename = infile2, use_iterators = True)
    sheet_ranges2 = wbFile2['Sheet1']

    #Header Population
    date = sheet_ranges1['J3'].value
    suffix = date.replace("-","")
    outputFile = open('C:\Users\Arshiya28\Desktop\MBA\Project\Software\Step-2\Output\CashFlowStatement_'+inputDate,"wb")
    outputFile.write("================================================================================================\n")
    outputFile.write("                                    CASH FLOW STATEMENT\n")
    outputFile.write("================================================================================================\n")
    outputFile.write("Company Name: "+sheet_ranges1['D3'].value+"\n")
    outputFile.write("For the Year Ending "+date+"\n")
    outputFile.write("================================================================================================\n")
    outputFile.write("CASH FLOW FROM OPERATING ACTIVIIES\n")
    outputFile.write("===================================\n")
    outputFile.write("      Net Income from Income Statement                                     $"+"%.2f" % netIncome + "\n")
    outputFile.write("      Additions to Cash\n")

    ws1 = wbFile1.get_sheet_by_name(name = 'Sheet1')
    ws2 = wbFile2.get_sheet_by_name(name = 'Sheet1')
    totalofOperations = 0

    totalofOperations = totalofOperations + netIncome
    # Depreciation
    p = 7
    q = ""
    depreciaition = 0
    for row in ws1.iter_rows('W7:W2000'):
        for cell in row:
            if str(cell.value) != "None":
                if str(cell.value) == "Depriciation":
                    q = "Y"+str(p)
                    if str(ws1.cell(q).value) != "None":
                        depreciaition = (ws1.cell(q).value)
                        outputFile.write("              Depriciation                                                 $"+"%.2f" % depreciaition + "\n")
                        totalofOperations = totalofOperations + depreciaition
                        break
        p = p + 1

    # Begining Cash
    p = 7
    q = ""
    for row in ws2.iter_rows('B7:B2000'):
        for cell in row:
            if str(cell.value) != "None":
                if str(cell.value) == "Cash_And_Cash_Equivalents":
                    q = "D"+str(p)
                    if str(ws2.cell(q).value) != "None":
                        beginingCash = (ws2.cell(q).value)
                        break
        p = p + 1

    # Accounts Receievable
    flag1 = 0
    flag2 = 0
    p = 7
    q = ""
    ar1 = 0
    for row in ws1.iter_rows('B7:B2000'):
        for cell in row:
            if str(cell.value) != "None":
                if str(cell.value) == "Accounts_Receivable":
                    flag1 = 1
                    q = "D"+str(p)
                    if str(ws1.cell(q).value) != "None":
                        ar1 = (ws1.cell(q).value)
                        break
        p = p + 1

    x = 7
    y = ""
    ar2 = 0
    for row in ws2.iter_rows('B7:B2000'):
        for cell in row:
            if str(cell.value) != "None":
                if str(cell.value) == "Accounts_Receivable":
                    flag2 = 1
                    y = "D"+str(x)
                    if str(ws2.cell(y).value) != "None":
                        ar2 = (ws2.cell(y).value)
        x = x + 1

    if flag1 == 1 and flag2 == 1:
        if ar1 > ar2 :
            substraction['Accounts Receivables'] = float(ar1-ar2)
        elif ar1 < ar2 :
            addition['Accounts Receivables'] = float(ar2-ar1)

    # Inventories
    flag1 = 0
    flag2 = 0
    p = 7
    q = ""
    i1 = 0
    for row in ws1.iter_rows('B7:B2000'):
        for cell in row:
            if str(cell.value) != "None":
                if str(cell.value) == "Inventories":
                    flag1 = 1
                    q = "D"+str(p)
                    if str(ws1.cell(q).value) != "None":
                        i1 = (ws1.cell(q).value)
                        break
        p = p + 1

    x = 7
    y = ""
    i2 = 0
    for row in ws2.iter_rows('B7:B2000'):
        for cell in row:
            if str(cell.value) != "None":
                if str(cell.value) == "Inventories":
                    flag2 = 1
                    y = "D"+str(x)
                    if str(ws2.cell(y).value) != "None":
                        i2 = (ws2.cell(y).value)
        x = x + 1

    if flag1 == 1 and flag2 == 1:
        if i1 > i2 :
            substraction['Inventories'] = float(i1-i2)
        elif i1 < i2 :
            addition['Inventories'] = float(i2-i1)

    # Prepaid Expenses
    flag1 = 0
    flag2 = 0
    p = 7
    q = ""
    pe1 = 0
    for row in ws1.iter_rows('B7:B2000'):
        for cell in row:
            if str(cell.value) != "None":
                if str(cell.value) == "Prepaid_Expenses":
                    flag1 = 1
                    q = "D"+str(p)
                    if str(ws1.cell(q).value) != "None":
                        pe1 = (ws1.cell(q).value)
                        break
        p = p + 1

    x = 7
    y = ""
    pe2 = 0
    for row in ws2.iter_rows('B7:B2000'):
        for cell in row:
            if str(cell.value) != "None":
                if str(cell.value) == "Prepaid_Expenses":
                    flag2 = 1
                    y = "D"+str(x)
                    if str(ws2.cell(y).value) != "None":
                        pe2 = (ws2.cell(y).value)
        x = x + 1

    if flag1 == 1 and flag2 == 1:
        if pe1 > pe2 :
            substraction['Prepaid Expenses'] = float(pe1-pe2)
        elif pe1 < pe2 :
            addition['Prepaid Expenses'] = float(pe2-pe1)


    # Deferred Taxes
    flag1 = 0
    flag2 = 0
    p = 7
    q = ""
    dit1 = 0
    for row in ws1.iter_rows('B7:B2000'):
        for cell in row:
            if str(cell.value) != "None":
                if str(cell.value) == "Deferred_Income_Taxes":
                    flag1 = 1
                    q = "D"+str(p)
                    if str(ws1.cell(q).value) != "None":
                        dit1 = (ws1.cell(q).value)
                        break
        p = p + 1

    x = 7
    y = ""
    dit2 = 0
    for row in ws2.iter_rows('B7:B2000'):
        for cell in row:
            if str(cell.value) != "None":
                if str(cell.value) == "Deferred_Income_Taxes":
                    flag2 = 1
                    y = "D"+str(x)
                    if str(ws2.cell(y).value) != "None":
                        dit2 = (ws2.cell(y).value)
        x = x + 1

    if flag1 == 1 and flag2 == 1:
        if dit1 > dit2 :
            addition['Deferred Taxes'] = float(dit1-dit2)
        elif dit1 < dit2 :
            substraction['Deferred Taxes'] = float(dit2-dit1)

    # Deferred Taxes
    flag1 = 0
    flag2 = 0
    p = 7
    q = ""
    dit1 = 0
    for row in ws1.iter_rows('B7:B2000'):
        for cell in row:
            if str(cell.value) != "None":
                if str(cell.value) == "Deferred_Income_Taxes":
                    flag1 = 1
                    q = "D"+str(p)
                    if str(ws1.cell(q).value) != "None":
                        dit1 = (ws1.cell(q).value)
                        break
        p = p + 1

    x = 7
    y = ""
    dit2 = 0
    for row in ws2.iter_rows('B7:B2000'):
        for cell in row:
            if str(cell.value) != "None":
                if str(cell.value) == "Deferred_Income_Taxes":
                    flag2 = 1
                    y = "D"+str(x)
                    if str(ws2.cell(y).value) != "None":
                        dit2 = (ws2.cell(y).value)
        x = x + 1

    if flag1 == 1 and flag2 == 1:
        if dit1 > dit2 :
            addition['Deferred Taxes'] = float(dit1-dit2)
        elif dit1 < dit2 :
            substraction['Deferred Taxes'] = float(dit2-dit1)

    # Account Payables
    flag1 = 0
    flag2 = 0
    p = 7
    q = ""
    ap1 = 0
    for row in ws1.iter_rows('H7:H2000'):
        for cell in row:
            if str(cell.value) != "None":
                if str(cell.value) == "Accounts_Payable_And_Accrued_Expenses":
                    flag1 = 1
                    q = "J"+str(p)
                    if str(ws1.cell(q).value) != "None":
                        ap1 = (ws1.cell(q).value)
                        break
        p = p + 1

    x = 7
    y = ""
    ap2 = 0
    for row in ws2.iter_rows('H7:H2000'):
        for cell in row:
            if str(cell.value) != "None":
                if str(cell.value) == "Accounts_Payable_And_Accrued_Expenses":
                    flag2 = 1
                    y = "J"+str(x)
                    if str(ws2.cell(y).value) != "None":
                        ap2 = (ws2.cell(y).value)
        x = x + 1

    if flag1 == 1 and flag2 == 1:
        if ap1 > ap2 :
            addition['Account Payables'] = float(ap1-ap2)
        elif ap1 < ap2 :
            substraction['Account Payables'] = float(ap2-ap1)

    # loss or gain on disposal
    p = 7
    q = ""
    cogs = 0
    flag1 = 0
    for row in ws1.iter_rows('R7:R2000'):
        for cell in row:
            if str(cell.value) != "None":
                if str(cell.value) == "Gain":
                    flag1 = 1
                    q = "T"+str(p)
                    if str(ws1.cell(q).value) != "None":
                        cogs = (ws1.cell(q).value)
                        substraction['Gain on Disposal'] = cogs
                        break
                elif str(cell.value) == "Loss":
                    flag1 = 1
                    q = "T"+str(p)
                    if str(ws1.cell(q).value) != "None":
                        cogs = (ws1.cell(q).value)
                        addition['Gain on Disposal'] = cogs
                        break
        p = p + 1


    # Displaying additions
    for category,amt in addition.items():
        line = category
        if len(line) < 61:
            trailing_spaces = 61-len(line)
            line = line + " " * trailing_spaces
        line = line + "$"+"%.2f" % float(amt)
        line = (line+"\n")
        outputFile.write("              "+line)
        totalofOperations =totalofOperations+amt
    outputFile.write("      Substractions from Cash\n")
    # Displaying substractions
    for category,amt in substraction.items():
        line = category
        if len(line) < 61:
            trailing_spaces = 61-len(line)
            line = line + " " * trailing_spaces
        line = line + "$"+"%.2f" % float(amt)
        line = (line+"\n")
        outputFile.write("              "+line)
        totalofOperations =totalofOperations-amt
    outputFile.write("------------------------------------------------------------------------------------------------\n")
    line = "NET CASH FLOW FROM OPERATING ACTIVITIES"
    if len(line) < 61:
        trailing_spaces = 61-len(line)
        line = line + " " * trailing_spaces
    line = line + "$"+"%.2f" % float(totalofOperations)
    outputFile.write(line+"\n")
    outputFile.write("------------------------------------------------------------------------------------------------\n")
    outputFile.write("CASH FLOW FROM INVESTING ACTIVIIES\n")
    outputFile.write("===================================\n")
    totalofInvesting = 0
    invadd = {}
    invsub = {}

    # Investing Activities in Assets
    flag1 = 0
    flag2 = 0
    p = 7
    q = ""
    pp1 = 0
    for row in ws1.iter_rows('B7:B2000'):
        for cell in row:
            if str(cell.value) != "None":
                if str(cell.value) == "Property, Plant,and_Equipment(net)":
                    flag1 = 1
                    q = "D"+str(p)
                    if str(ws1.cell(q).value) != "None":
                        pp1 = (ws1.cell(q).value)
                        break
        p = p + 1

    x = 7
    y = ""
    pp2 = 0
    for row in ws2.iter_rows('B7:B2000'):
        for cell in row:
            if str(cell.value) != "None":
                if str(cell.value) == "Property, Plant,and_Equipment(net)":
                    flag2 = 1
                    y = "D"+str(x)
                    if str(ws2.cell(y).value) != "None":
                        pp2 = (ws2.cell(y).value)
        x = x + 1

    if flag1 == 1 and flag2 == 1:
        if pp1 > pp2 :
            invsub['Increase in Property, Plant,and_Equipment(net)'] = float(pp1-pp2)
        elif pp1 < pp2 :
            invadd['Decrease in Property, Plant,and_Equipment(net)'] = float(pp2-pp1)

    # Building
    flag1 = 0
    flag2 = 0
    p = 7
    q = ""
    b1 = 0
    for row in ws1.iter_rows('B7:B2000'):
        for cell in row:
            if str(cell.value) != "None":
                if str(cell.value) == "Building":
                    flag1 = 1
                    q = "D"+str(p)
                    if str(ws1.cell(q).value) != "None":
                        b1 = (ws1.cell(q).value)
                        break
        p = p + 1

    x = 7
    y = ""
    b2 = 0
    for row in ws2.iter_rows('B7:B2000'):
        for cell in row:
            if str(cell.value) != "None":
                if str(cell.value) == "Building":
                    flag2 = 1
                    y = "D"+str(x)
                    if str(ws2.cell(y).value) != "None":
                        b2 = (ws2.cell(y).value)
        x = x + 1

    if flag1 == 1 and flag2 == 1:
        if b1 > b2 :
            invadd['Increase in Building'] = float(b1-b2)
        elif b1 < b2 :
            invsub['Decrease in Building'] = float(b2-b1)
    elif flag1 == 1 and flag2 == 0:
        invsub['Increase in Building'] = float(b1)

    # Equity Investments
    flag1 = 0
    flag2 = 0
    p = 7
    q = ""
    ei1 = 0
    for row in ws1.iter_rows('B7:B2000'):
        for cell in row:
            if str(cell.value) != "None":
                if str(cell.value) == "Equity_Investments":
                    flag1 = 1
                    q = "D"+str(p)
                    if str(ws1.cell(q).value) != "None":
                        ei1 = (ws1.cell(q).value)
                        break
        p = p + 1

    x = 7
    y = ""
    ei2 = 0
    for row in ws2.iter_rows('B7:B2000'):
        for cell in row:
            if str(cell.value) != "None":
                if str(cell.value) == "Equity_Investments":
                    flag2 = 1
                    y = "D"+str(x)
                    if str(ws2.cell(y).value) != "None":
                        ei2 = (ws2.cell(y).value)
                        break
        x = x + 1

    if flag1 == 1 and flag2 == 1:
        if ei1 > ei2 :
            invsub['Equity Investment'] = float(ei1-ei2)
        elif ei1 < ei2 :
            invadd['Equity Investments'] = float(ei2-ei1)

    # Displaying additions
    for category,amt in invadd.items():
        line = category
        if len(line) < 61:
            trailing_spaces = 61-len(line)
            line = line + " " * trailing_spaces
        line = line + "$"+"%.2f" % float(amt)
        line = (line+"\n")
        outputFile.write("              "+line)
        totalofInvesting =totalofInvesting+amt
    # Displaying substractions
    for category,amt in invsub.items():
        line = category
        if len(line) < 61:
            trailing_spaces = 61-len(line)
            line = line + " " * trailing_spaces
        line = line + "$"+"%.2f" % float(amt)
        line = (line+"\n")
        outputFile.write("              "+line)
        totalofInvesting =totalofInvesting-amt
    outputFile.write("------------------------------------------------------------------------------------------------\n")
    line = "NET CASH FLOW FROM INVESTING ACTIVITIES"
    if len(line) < 61:
        trailing_spaces = 61-len(line)
        line = line + " " * trailing_spaces
    line = line + "$"+"%.2f" % float(totalofInvesting)
    outputFile.write(line+"\n")
    outputFile.write("------------------------------------------------------------------------------------------------\n")
    outputFile.write("CASH FLOW FROM FINANCING ACTIVIIES\n")
    outputFile.write("===================================\n")
    totalofFinancing = 0
    finadd = {}
    finsub = {}

    # Common Stock or Capital
    flag1 = 0
    flag2 = 0
    p = 7
    q = ""
    c1 = 0
    for row in ws1.iter_rows('M7:M2000'):
        for cell in row:
            if str(cell.value) != "None":
                if str(cell.value) == "Capital":
                    flag1 = 1
                    q = "O"+str(p)
                    if str(ws1.cell(q).value) != "None":
                        c1 = (ws1.cell(q).value)
                        break
        p = p + 1

    x = 7
    y = ""
    c2 = 0
    for row in ws2.iter_rows('M7:M2000'):
        for cell in row:
            if str(cell.value) != "None":
                if str(cell.value) == "Capital":
                    flag2 = 1
                    y = "O"+str(x)
                    if str(ws2.cell(y).value) != "None":
                        c2 = (ws2.cell(y).value)
        x = x + 1

    if flag1 == 1 and flag2 == 1:
        if c1 > c2 :
            finadd['Increase in Common Stock/Capital'] = float(c1-c2)
        elif c1 < c2 :
            finsub['Decrease in Common Stock/Capital'] = float(c2-c1)

    # Note Payables
    flag1 = 0
    flag2 = 0
    p = 7
    q = ""
    np1 = 0
    for row in ws1.iter_rows('H7:H2000'):
        for cell in row:
            if str(cell.value) != "None":
                if str(cell.value) == "Note_Payable":
                    flag1 = 1
                    q = "J"+str(p)
                    if str(ws1.cell(q).value) != "None":
                        np1 = (ws1.cell(q).value)
                        break
        p = p + 1

    x = 7
    y = ""
    np2 = 0
    for row in ws2.iter_rows('H7:H2000'):
        for cell in row:
            if str(cell.value) != "None":
                if str(cell.value) == "Note_Payable":
                    flag2 = 1
                    y = "J"+str(x)
                    if str(ws2.cell(y).value) != "None":
                        np2 = (ws2.cell(y).value)
        x = x + 1

    if flag1 == 1 and flag2 == 1:
        if np1 > np2 :
            finadd['Note Payables'] = float(np1-np2)
        elif np1 < np2 :
            finsub['Note Payables'] = float(np2-np1)

    # Note Payables
    flag1 = 0
    flag2 = 0
    p = 7
    q = ""
    bp1 = 0
    for row in ws1.iter_rows('H7:H2000'):
        for cell in row:
            if str(cell.value) != "None":
                if str(cell.value) == "Bonds_Payable":
                    flag1 = 1
                    q = "J"+str(p)
                    if str(ws1.cell(q).value) != "None":
                        bp1 = (ws1.cell(q).value)
                        break
        p = p + 1

    x = 7
    y = ""
    bp2 = 0
    for row in ws2.iter_rows('H7:H2000'):
        for cell in row:
            if str(cell.value) != "None":
                if str(cell.value) == "Bonds_Payable":
                    flag2 = 1
                    y = "J"+str(x)
                    if str(ws2.cell(y).value) != "None":
                        bp2 = (ws2.cell(y).value)
        x = x + 1

    if flag1 == 1 and flag2 == 1:
        if bp1 > bp2 :
            finadd['Bonds Payables'] = float(bp1-bp2)
        elif bp1 < bp2 :
            finsub['Bonds Payables'] = float(bp2-bp1)

    # Note Payables
    flag1 = 0
    flag2 = 0
    p = 7
    q = ""
    lt1 = 0
    for row in ws1.iter_rows('H7:H2000'):
        for cell in row:
            if str(cell.value) != "None":
                if str(cell.value) == "Long-term_debt":
                    flag1 = 1
                    q = "J"+str(p)
                    if str(ws1.cell(q).value) != "None":
                        lt1 = (ws1.cell(q).value)
                        break
        p = p + 1

    x = 7
    y = ""
    lt2 = 0
    for row in ws2.iter_rows('H7:H2000'):
        for cell in row:
            if str(cell.value) != "None":
                if str(cell.value) == "Long-term_debt":
                    flag2 = 1
                    y = "J"+str(x)
                    if str(ws2.cell(y).value) != "None":
                        lt2 = (ws2.cell(y).value)
        x = x + 1

    if flag1 == 1 and flag2 == 1:
        if lt1 > lt2 :
            finadd['Long-term debt'] = float(lt1-lt2)
        elif lt1 < lt2 :
            finsub['Long-term debt'] = float(lt2-lt1)

    # Mortgage Payables
    flag1 = 0
    flag2 = 0
    p = 7
    q = ""
    mp1 = 0
    for row in ws1.iter_rows('H7:H2000'):
        for cell in row:
            if str(cell.value) != "None":
                if str(cell.value) == "Mortgage_Payable":
                    flag1 = 1
                    q = "J"+str(p)
                    if str(ws1.cell(q).value) != "None":
                        mp1 = (ws1.cell(q).value)
                        break
        p = p + 1

    x = 7
    y = ""
    mp2 = 0
    for row in ws2.iter_rows('H7:H2000'):
        for cell in row:
            if str(cell.value) != "None":
                if str(cell.value) == "Mortgage_Payable":
                    flag2 = 1
                    y = "J"+str(x)
                    if str(ws2.cell(y).value) != "None":
                        mp2 = (ws2.cell(y).value)
        x = x + 1

    if flag1 == 1 and flag2 == 1:
        if mp1 > mp2 :
            finadd['Mortgage Payable'] = float(mp1-mp2)
        elif mp1 < mp2 :
            finsub['Mortgage Payable'] = float(mp2-mp1)

    # Displaying additions
    for category,amt in finadd.items():
        line = category
        if len(line) < 61:
            trailing_spaces = 61-len(line)
            line = line + " " * trailing_spaces
        line = line + "$"+"%.2f" % float(amt)
        line = (line+"\n")
        outputFile.write("              "+line)
        totalofFinancing =totalofFinancing+amt
    # Displaying substractions
    for category,amt in finsub.items():
        line = category
        if len(line) < 61:
            trailing_spaces = 61-len(line)
            line = line + " " * trailing_spaces
        line = line + "$"+"%.2f" % float(amt)
        line = (line+"\n")
        outputFile.write("              "+line)
        totalofFinancing =totalofFinancing-amt
    totalofFinancing = totalofFinancing - dividends
    outputFile.write("              Dividends                                                    $%.2f\n" % float(dividends))
    outputFile.write("------------------------------------------------------------------------------------------------\n")
    line = "NET CASH FLOW FROM FINANCING ACTIVITIES"
    if len(line) < 61:
        trailing_spaces = 61-len(line)
        line = line + " " * trailing_spaces
    line = line + "$"+"%.2f" % float(totalofFinancing)
    outputFile.write(line+"\n")
    outputFile.write("------------------------------------------------------------------------------------------------\n")
    totalCash = 0
    totalCash = totalCash + totalofOperations + totalofInvesting + totalofFinancing
    line = "NET CASH CHANGE DURING THIS FINACIAL YEAR"
    if len(line) < 61:
        trailing_spaces = 61-len(line)
        line = line + " " * trailing_spaces
    line = line + "$"+"%.2f" % float(totalCash)
    outputFile.write(line+"\n")

    line = "BEGINING CASH BALANCE "
    if len(line) < 61:
        trailing_spaces = 61-len(line)
        line = line + " " * trailing_spaces
    line = line + "$"+"%.2f" % float(beginingCash)
    outputFile.write(line+"\n")
    totalremaining = totalCash + beginingCash
    outputFile.write("------------------------------------------------------------------------------------------------\n\n")
    outputFile.write("================================================================================================\n")
    line = "ENDING CASH BALANCE "
    if len(line) < 61:
        trailing_spaces = 61-len(line)
        line = line + " " * trailing_spaces
    line = line + "$"+"%.2f" % float(totalremaining)
    outputFile.write(line+"\n")
    outputFile.write("================================================================================================\n")
    outputFile.close()


def createRetainedEarningStatement(inFile1,netIncome,dividends,inputDate):
    # First Balance Sheet
    wbFile1 =  load_workbook(filename = inFile1, use_iterators = True)
    sheet_ranges1 = wbFile1['Sheet1']
    ws1 = wbFile1.get_sheet_by_name(name = 'Sheet1')
    # Previous Retained Earning
    p = 7
    q = ""
    retear = 0
    flag = 0
    totals = 0
    for row in ws1.iter_rows('M7:M2000'):
        for cell in row:
            if str(cell.value) != "None":
                if str(cell.value) == "Pervious_Retained_Earnings":
                    flag = 1
                    q = "O"+str(p)
                    if str(ws1.cell(q).value) != "None":
                        retear = (ws1.cell(q).value)
        p = p + 1

    #Header Population
    date = sheet_ranges1['J3'].value
    suffix = date.replace("-","")
    outputFile = open('C:\Users\Arshiya28\Desktop\MBA\Project\Software\Step-2\Output\RetainedEarningStatement_'+inputDate,"wb")
    outputFile.write("================================================================================================\n")
    outputFile.write("                                    RETAINED EARNING STATEMENT\n")
    outputFile.write("================================================================================================\n")
    outputFile.write("Company Name: "+sheet_ranges1['D3'].value+"\n")
    outputFile.write("For the Year Ending "+date+"\n")
    outputFile.write("================================================================================================\n")
    outputFile.write("      Previous Retained Earnings                                       $"+"%.2f" % retear + "\n")
    outputFile.write("      Net Income from Income Statement                                 $"+"%.2f" % netIncome + "\n")
    outputFile.write("      Dividends paid to Shareholders                        $"+"%.2f" % dividends + "\n")
    outputFile.write("------------------------------------------------------------------------------------------------\n")
    totals = (retear + netIncome) - dividends
    outputFile.write("      RETAINED EARNING at "+date+"                                $"+"%.2f" % totals +"\n")
    outputFile.write("------------------------------------------------------------------------------------------------\n")
    outputFile.close()




# Main Program
if __name__ == "__main__":
    print "Program to Generate Financial Statements."
    print "------------------------------------------------------------------------------------------------------"
    inputDate = raw_input("Enter the date for which financial statements to be generated(ddmmyyyy): \t")
    balanceSheet = 'C:\Users\Arshiya28\Desktop\MBA\Project\Software\Step-2\Input\Balance_Sheet_'+inputDate+'.xlsm'
    flag = 0
    netIncome = 0
    if not os.path.isfile(balanceSheet) :
        print "\nInput Balance Sheet for the date "+inputDate+" does not exists!!!!"
        print "Exiting the program."
        sys.exit(99)

    while True:
        print "============================================================="
        print "Financial Statements:"
        print "============================================================="
        print "1. Income Statement"
        print "2. Cash Flow Statement"
        print "3. Retained Earning Statement"
        print "4. Exit"
        option=raw_input("Enter the Option: \t")
        if option not in ["1","2", "3", "4"]:
            print "Invalid Input!!!!"
            print "Enter an option 1 to 4"
        elif option == "1":
            print "Generating Income Statement!!!!!"
            flag = 1
            netIncome = createIncomeStatement(balanceSheet,inputDate)
            if os.path.isfile('C:\Users\Arshiya28\Desktop\MBA\Project\Software\Step-2\Output\IncomeStatement_'+inputDate):
                print "Income Statement Generated."
                print "Income Statement is placed in below location."
                print 'C:\Users\Arshiya28\Desktop\MBA\Project\Software\Step-2\Output\IncomeStatement_'+inputDate
        elif option == "2":
            if not os.path.isfile('C:\Users\Arshiya28\Desktop\MBA\Project\Software\Step-2\Output\IncomeStatement_'+inputDate):
                print "For generating cash flow statement. Income statement should be generated."
                print "Please Select option 1 before selecting option 2."
            else:
                fp = open('C:\Users\Arshiya28\Desktop\MBA\Project\Software\Step-2\Output\IncomeStatement_'+inputDate,"r")
                contents = []
                contents = fp.readlines()
                netIncome = contents[-2].split("$")[1].strip()

                x = int(inputDate[-4:])-1
                prevBalDate = inputDate[:4]+str(x)
                prevBalanceSheet = 'C:\Users\Arshiya28\Desktop\MBA\Project\Software\Step-2\Input\Balance_Sheet_'+prevBalDate+'.xlsm'
                if not os.path.isfile(prevBalanceSheet):
                    print "Prev Year Balance Sheet Dated:" + prevBalDate +" does not exists!!!"
                    infile2 = raw_input("Enter the absolute path of the balance sheet to be considered.")
                    if not os.path.isfile(infile2):
                        print infile2 + " file does not exists!!!!!!"
                        file2 = ""
                    else:
                        file2 = infile2
                else:
                    file2=prevBalanceSheet
                if file2 != "":
                    dividends = raw_input("Enter the dividends issued to the shareholders: \t")
                    if not dividends.isdigit():
                        if dividends.find(".") != -1:
                            dotSplit = []
                            dotSplit = dividends.split(".")
                            if len(dotSplit) == 2:
                                if dotSplit[0].isdigit() and dotSplit[1].isdigit():
                                    flag3=1
                                else:
                                    flag3=0
                            else:
                                flag3 = 0
                        else:
                            flag3=0
                    else:
                        flag3 =1
                    if flag3 == 1:
                        print "Generating Cash Flow Statement!!!!!"
                        createCashFlowStatements(balanceSheet,file2,float(netIncome),float(dividends),inputDate)
                        if os.path.isfile('C:\Users\Arshiya28\Desktop\MBA\Project\Software\Step-2\Output\CashFlowStatement_'+inputDate):
                            print "Cash Flow Statement Generated."
                            print "Cash Flow Statement is placed in below location."
                            print 'C:\Users\Arshiya28\Desktop\MBA\Project\Software\Step-2\Output\CashFlowStatement_'+inputDate
                    else:
                        print "Invalid Dividends amount entered."
                        print "Enter a valid number"
        elif option == "3":
            if not os.path.isfile('C:\Users\Arshiya28\Desktop\MBA\Project\Software\Step-2\Output\IncomeStatement_'+inputDate):
                print "For generating cash flow statement. Income statement should be generated."
                print "Please Select option 1 before selecting option 3."
            else:
                fp = open('C:\Users\Arshiya28\Desktop\MBA\Project\Software\Step-2\Output\IncomeStatement_'+inputDate,"r")
                contents = fp.readlines()
                netIncome = contents[-2].split("$")[1].strip()
                dividends = raw_input("Enter the dividends issued to the shareholders: \t")
                if not dividends.isdigit():
                    if dividends.find(".") != -1:
                        dotSplit = []
                        dotSplit = dividends.split(".")
                        if len(dotSplit) == 2:
                            if dotSplit[0].isdigit() and dotSplit[1].isdigit():
                                flag3=1
                            else:
                                flag3=0
                        else:
                            flag3 = 0
                    else:
                        flag3=0
                else:
                    flag3 =1
                if flag3 == 1:
                    print "Generating Retained Earning Statement!!!!!"
                    createRetainedEarningStatement(balanceSheet,float(netIncome),float(dividends),inputDate)
                    if os.path.isfile('C:\Users\Arshiya28\Desktop\MBA\Project\Software\Step-2\Output\RetainedEarningStatement_'+inputDate):
                        print "Retained Earning Statement Generated."
                        print "Retained Earning Statement is placed in below location."
                        print 'C:\Users\Arshiya28\Desktop\MBA\Project\Software\Step-2\Output\RetainedEarningStatement_'+inputDate
                else:
                    print "Invalid Dividends amount entered."
                    print "Enter a valid number"
        elif option == "4":
            print "Option 4 Selected !!!!!!!!!!!"
            print "Existing the Program."
            sys.exit(0)
